import csv
import json
from collections import defaultdict
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _to_float(v) -> Optional[float]:
    try:
        return float(v) if v is not None else None
    except Exception:
        return None


def _offer_note(row: Dict[str, Any]) -> str:
    unit_type = (row.get("unit_type") or "").lower()
    seal_status = (row.get("seal_status") or "").lower()
    variant = row.get("variant_name") or ""

    if unit_type == "case":
        base = "sealed case"
    elif seal_status == "no_shrink":
        base = "no shrink"
    else:
        base = "sealed"

    return f"{base} | {variant}" if variant else base


def _effective_ppp(row: Dict[str, Any]) -> Optional[float]:
    return _to_float(row.get("landed_price_per_pack") or row.get("price_per_pack"))


# ----------------------------------------------------------
# RAW OUTPUTS
# ----------------------------------------------------------

def write_csv(path: str, rows: List[Dict[str, Any]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_json(path: str, rows: List[Dict[str, Any]]) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)


# ----------------------------------------------------------
# COVERAGE + NEAR MISSES (tuning files)
# ----------------------------------------------------------

def write_coverage_csv(path: str, coverage_rows: List[Dict[str, Any]]) -> None:
    headers = [
        "site_name",
        "discovered_urls",
        "processed_urls",
        "matched_rows",
        "keyword_hits_but_filtered",
        "unknown_product_type",
    ]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in coverage_rows:
            w.writerow({k: r.get(k) for k in headers})


def write_near_misses_csv(path: str, rows: List[Dict[str, Any]]) -> None:
    headers = ["site_name", "title", "reason", "url"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k) for k in headers})


# ----------------------------------------------------------
# BEST BY SET (CSV + XLSX)
# ----------------------------------------------------------

def build_best_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    groups = defaultdict(list)

    for r in rows:
        key = (r.get("franchise"), r.get("set_name"), r.get("product_type"))
        groups[key].append(r)

    best_rows: List[Dict[str, Any]] = []

    for (franchise, set_name, product_type), items in groups.items():
        candidates = [x for x in items if x.get("in_stock") and _effective_ppp(x) is not None]
        if not candidates:
            continue

        candidates.sort(key=lambda x: _effective_ppp(x) or 10**9)

        best = candidates[0]
        second = candidates[1] if len(candidates) > 1 else None

        best_rows.append({
            "timestamp": best.get("timestamp"),
            "franchise": franchise,
            "set_name": set_name,
            "product_type": product_type,
            "best_site": best.get("site_name"),
            "best_offer_note": _offer_note(best),
            "best_url": best.get("sku_or_url"),
            "best_price": best.get("price"),
            "best_price_per_pack": best.get("price_per_pack"),
            "best_landed_price": best.get("landed_price"),
            "best_landed_price_per_pack": best.get("landed_price_per_pack"),
            "best_effective_price_per_pack": _effective_ppp(best),
            "second_best_effective_price_per_pack": _effective_ppp(second) if second else None,
            "num_in_stock_offers": len(candidates),
        })

    best_rows.sort(key=lambda r: r.get("best_effective_price_per_pack") or 10**9)
    return best_rows


def write_best_by_set_csv(path: str, rows: List[Dict[str, Any]]) -> None:
    best_rows = build_best_rows(rows)
    if not best_rows:
        return

    headers = list(best_rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in best_rows:
            w.writerow(r)


def write_best_by_set_xlsx(path: str, rows: List[Dict[str, Any]]) -> None:
    best_rows = build_best_rows(rows)
    if not best_rows:
        return

    wb = Workbook()
    wb.remove(wb.active)

    headers = list(best_rows[0].keys())

    def create_sheet(sheet_name: str, franchise_name: str):
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)

        for r in best_rows:
            if (r.get("franchise") or "") == franchise_name:
                ws.append([r.get(h) for h in headers])

        for i, col in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, min(55, len(col) + 2))

        ws.freeze_panes = "A2"

    create_sheet("Pokemon", "Pokemon")
    create_sheet("One Piece", "One Piece")

    wb.save(path)
